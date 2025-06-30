#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
from datetime import datetime
import os

class ExcelExporter:
    """엑셀 파일 내보내기 클래스"""
    
    def __init__(self):
        self.column_order = [
            '매물명',
            '가격',
            '가격표시',
            '면적',
            '면적표시',
            '층수',
            '방향',
            '주소',
            'URL',
            '수집시간'
        ]
    
    def export_to_excel(self, data, filepath):
        """데이터를 엑셀 파일로 내보내기"""
        try:
            if not data:
                print("⚠️ 내보낼 데이터가 없습니다.")
                return False
            
            # DataFrame 생성
            df = pd.DataFrame(data)
            
            # 컬럼 순서 정리
            df = self._organize_columns(df)
            
            # 엑셀 파일로 저장
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # 메인 데이터 시트
                df.to_excel(writer, sheet_name='매물정보', index=False)
                
                # 통계 정보 시트
                self._create_summary_sheet(writer, df)
                
                # 필터링 가이드 시트
                self._create_filter_guide_sheet(writer)
            
            print(f"✅ 엑셀 파일이 성공적으로 저장되었습니다: {filepath}")
            return True
            
        except Exception as e:
            print(f"❌ 엑셀 파일 저장 중 오류 발생: {str(e)}")
            return False
    
    def _organize_columns(self, df):
        """컬럼 순서 정리"""
        # 존재하는 컬럼만 선택
        existing_columns = [col for col in self.column_order if col in df.columns]
        
        # 기타 컬럼들 추가
        other_columns = [col for col in df.columns if col not in existing_columns]
        final_columns = existing_columns + other_columns
        
        return df[final_columns]
    
    def _create_summary_sheet(self, writer, df):
        """통계 정보 시트 생성"""
        try:
            summary_data = []
            
            # 기본 통계
            summary_data.append(['총 매물 수', len(df)])
            
            # 가격 통계
            if '가격' in df.columns:
                prices = df['가격'].dropna()
                if len(prices) > 0:
                    summary_data.append(['평균 가격 (만원)', f"{prices.mean():,.0f}"])
                    summary_data.append(['최저 가격 (만원)', f"{prices.min():,.0f}"])
                    summary_data.append(['최고 가격 (만원)', f"{prices.max():,.0f}"])
                    summary_data.append(['가격 중간값 (만원)', f"{prices.median():,.0f}"])
            
            # 면적 통계
            if '면적' in df.columns:
                areas = df['면적'].dropna()
                if len(areas) > 0:
                    summary_data.append(['평균 면적 (㎡)', f"{areas.mean():.1f}"])
                    summary_data.append(['최소 면적 (㎡)', f"{areas.min():.1f}"])
                    summary_data.append(['최대 면적 (㎡)', f"{areas.max():.1f}"])
            
            # 층수 분포
            if '층수' in df.columns:
                floor_counts = df['층수'].value_counts()
                for floor, count in floor_counts.head(10).items():
                    summary_data.append([f'{floor}층', count])
            
            # 방향 분포
            if '방향' in df.columns:
                direction_counts = df['방향'].value_counts()
                for direction, count in direction_counts.head(10).items():
                    summary_data.append([f'{direction}향', count])
            
            # 수집 시간 정보
            if '수집시간' in df.columns:
                summary_data.append(['수집 시작 시간', df['수집시간'].min()])
                summary_data.append(['수집 종료 시간', df['수집시간'].max()])
            
            # 통계 데이터프레임 생성
            summary_df = pd.DataFrame(summary_data, columns=['항목', '값'])
            summary_df.to_excel(writer, sheet_name='통계정보', index=False)
            
        except Exception as e:
            print(f"통계 시트 생성 중 오류: {str(e)}")
    
    def _create_filter_guide_sheet(self, writer):
        """필터링 가이드 시트 생성"""
        try:
            guide_data = [
                ['필터링 방법', '설명'],
                ['가격 필터', '=FILTER(매물정보!A:J, 매물정보!B:B>=10000)'],
                ['면적 필터', '=FILTER(매물정보!A:J, 매물정보!D:D>=20)'],
                ['특정 지역 필터', '=FILTER(매물정보!A:J, ISNUMBER(SEARCH("강남", 매물정보!H:H)))'],
                ['최신 매물 필터', '=SORT(매물정보!A:J, 매물정보!J:J, -1)'],
                ['가격순 정렬', '=SORT(매물정보!A:J, 매물정보!B:B, 1)'],
                ['면적순 정렬', '=SORT(매물정보!A:J, 매물정보!D:D, 1)'],
                ['', ''],
                ['사용 팁', ''],
                ['1. 가격 필터', '원하는 가격대의 매물만 보기'],
                ['2. 면적 필터', '원하는 면적의 매물만 보기'],
                ['3. 지역 필터', '특정 지역의 매물만 보기'],
                ['4. 정렬 기능', '가격, 면적, 수집시간 순으로 정렬'],
                ['5. 중복 제거', '=UNIQUE(매물정보!A:J)'],
            ]
            
            guide_df = pd.DataFrame(guide_data, columns=['필터링 방법', '설명'])
            guide_df.to_excel(writer, sheet_name='필터가이드', index=False)
            
        except Exception as e:
            print(f"필터 가이드 시트 생성 중 오류: {str(e)}")
    
    def export_with_formatting(self, data, filepath):
        """서식이 적용된 엑셀 파일로 내보내기"""
        try:
            if not data:
                return False
            
            # 기본 엑셀 파일 생성
            success = self.export_to_excel(data, filepath)
            
            if success:
                # 서식 적용
                self._apply_formatting(filepath)
            
            return success
            
        except Exception as e:
            print(f"서식 적용 중 오류: {str(e)}")
            return False
    
    def _apply_formatting(self, filepath):
        """엑셀 파일에 서식 적용"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # 워크북 로드
            wb = load_workbook(filepath)
            ws = wb['매물정보']
            
            # 헤더 서식
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 헤더에 서식 적용
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 컬럼 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 테두리 설정
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
            
            # 워크북 저장
            wb.save(filepath)
            print("✅ 엑셀 서식이 적용되었습니다.")
            
        except Exception as e:
            print(f"서식 적용 중 오류: {str(e)}") 